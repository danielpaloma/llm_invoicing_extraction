import pandas as pd
from openpyxl import load_workbook

def save_json_to_excel(json_obj, filename, excel_path, invoices_sheet, line_items_sheet):
    """
    Appends invoice and line items data from a JSON object to an Excel file.
    Sheet names are provided as arguments.
    Only new data is written, not the entire sheet.
    """
    invoice_data = {
        "filename": filename,
        "invoice_number": json_obj.get("invoice_number"),
        "vendor": json_obj.get("vendor"),
        "recipient": json_obj.get("recipient"),
        "invoice_date": json_obj.get("invoice_date"),
        "invoice_total": json_obj.get("invoice_total")
    }
    invoice_df = pd.DataFrame([invoice_data])

    line_items = json_obj.get("line_items", [])
    line_items_df = pd.DataFrame([
        {
            "filename": filename,
            "invoice_number": json_obj.get("invoice_number"),
            "description": item.get("description"),
            "quantity": item.get("quantity"),
            "rate": item.get("rate"),
            "total": item.get("total")
        }
        for item in line_items
    ])

    # If file exists, append without loading all data
    if pd.io.common.file_exists(excel_path):
        book = load_workbook(excel_path)
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Append to invoices_sheet
            print("Appending to output Excel file...")
            invoice_df.to_excel(writer, sheet_name=invoices_sheet, index=False, header=False, startrow=book[invoices_sheet].max_row)
            # Append to line_items_sheet
            line_items_df.to_excel(writer, sheet_name=line_items_sheet, index=False, header=False, startrow=book[line_items_sheet].max_row)
    else:
        # Create new file with headers
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            invoice_df.to_excel(writer, sheet_name=invoices_sheet, index=False)
            line_items_df.to_excel(writer, sheet_name=line_items_sheet, index=False)